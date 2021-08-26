import re
import openpyxl
import getpass
import pandas as pd

def snomed_mixed_to_term (snomed_mixed_string):
    snomed_term_string = re.sub("\d+", " ", snomed_mixed_string)
    snomed_term_string = re.sub('\|', '', snomed_term_string)
    #print("This is a term expression: " + snomed_numeric_string)
    return snomed_term_string


def snomed_mixed_to_numeric (snomed_mixed_string):
    snomed_numeric_string = re.sub(r"[a-zA-Z]" , '', snomed_mixed_string)
    snomed_numeric_string = re.sub(' ', '', snomed_numeric_string)
    snomed_numeric_string = re.sub('\|', '', snomed_numeric_string)
    snomed_numeric_string = re.sub('/', '', snomed_numeric_string)
    snomed_numeric_string = re.sub('-', '', snomed_numeric_string)

    #for "(" and ")" handle case where used for nested grouping in post-coordination
    snomed_numeric_string = re.sub('\([a-zA-Z]\)', '', snomed_numeric_string)
    snomed_numeric_string = re.sub('\(\)', '', snomed_numeric_string)
    #print("This is a term expression: " + snomed_term_string)
    return snomed_numeric_string

def brackets_intact(s):
    if s.endswith(")"):
        return True
    else:
        return False


def excel_input_output (one_column_xls):
    wb = openpyxl.load_workbook(one_column_xls)
    ws = wb.active
    rownum = 1
    for row in ws.values:
        for value in row:
            str_value = str(value)
            #print(value)
            numeric = snomed_mixed_to_numeric(str_value)
            num_cellref = ws.cell(row= rownum, column=2)
            num_cellref.value = numeric
            print(numeric)
            term = snomed_mixed_to_term(str_value)
            term_cellref = ws.cell(row=rownum, column=3)
            term_cellref.value = term
            #todo fix the following
            # if not brackets_intact(term):
            #     message_cellref = ws.cell(row=rownum, column=4)
            #     message_cellref.value = "There is something wrong here"
        rownum += 1
    wb.save(r"C:\Users\france\Desktop\mio42\worksheet\output-regex.xlsx")
    print("saved to folder")

if __name__ == "__main__":
    excel_input_output(r"C:\Users\france\Desktop\mio42\worksheet\input-regex.xlsx")
    snomed_mixed_string = ("363787002 |Observable entity (observable entity)|: 704321009 |Characterizes (attribute)| = 223458004 | Informing (procedure) |, 370131001 |Recipient category (attribute)|= 35359004 |Family (social concept)|")
    term = snomed_mixed_to_term(snomed_mixed_string)
    numeric = snomed_mixed_to_numeric(snomed_mixed_string)
    print(term)
    print(numeric)
